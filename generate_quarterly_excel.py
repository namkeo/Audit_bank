"""
Generate a clean, formatted Excel file for quarterly risk scores results.
Creates a single, easy-to-read Excel file with formatting and filters.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

def create_quarterly_results_excel():
    """Create a formatted Excel file for quarterly risk scores."""
    
    try:
        # Load the quarterly scores data
        df = pd.read_csv('outputs/quarterly_risk_scores.csv')
        
        logger.info("📊 Creating Quarterly Results Excel File...")
        logger.info(f"   Total Records: {len(df)}")
        logger.info(f"   Banks: {df['bank_name'].nunique()}")
        logger.info(f"   Quarters: {df['period'].nunique()}")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Quarterly Results"
        
        # Add title
        ws['A1'] = "QUARTERLY RISK SCORES REPORT"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25
        
        # Add metadata
        ws['A2'] = f"Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=9, italic=True, color="666666")
        ws['A3'] = f"Period: {df['period'].min()} to {df['period'].max()}"
        ws['A3'].font = Font(size=9, italic=True, color="666666")
        
        # Add headers (starting row 5)
        headers = ['Period', 'Bank Name', 'Credit Score', 'Liquidity Score', 
                   'Anomaly Score', 'Quarterly Risk Score', 'Risk Level']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Define risk level colors
        risk_colors = {
            'MINIMAL': 'C6EFCE',  # Light green
            'LOW': 'FFEB9C',      # Light yellow
            'MEDIUM': 'FFC7CE',   # Light orange/pink
            'HIGH': 'FF0000',     # Red
            'CRITICAL': '660000'  # Dark red
        }
        
        # Add data rows
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row_num, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 6):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal="center" if col_num >= 3 else "left")
                
                # Color code risk levels
                if col_num == 7:  # Risk Level column
                    risk_level = value
                    if risk_level in risk_colors:
                        text_color = "FFFFFF" if risk_level in ['HIGH', 'CRITICAL'] else "000000"
                        cell.fill = PatternFill(start_color=risk_colors[risk_level], 
                                              end_color=risk_colors[risk_level], 
                                              fill_type="solid")
                        cell.font = Font(bold=True, color=text_color)
                
                # Format score columns (3-6)
                if col_num >= 3 and col_num <= 6:
                    if isinstance(value, (int, float)):
                        cell.value = round(value, 2)
                        cell.number_format = '0.00'
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 12
        
        # Add autofilter
        ws.auto_filter.ref = f"A5:G{5 + len(df)}"
        
        # Freeze header rows
        ws.freeze_panes = 'A6'
        
        # Save file
        output_file = 'outputs/quarterly_results.xlsx'
        wb.save(output_file)
        
        logger.info(f"✅ Successfully created: {output_file}")
        logger.info(f"   Rows: {len(df)} quarterly records")
        logger.info(f"   Features:")
        logger.info(f"   ✓ Color-coded risk levels")
        logger.info(f"   ✓ Auto-filter enabled")
        logger.info(f"   ✓ Frozen headers")
        logger.info(f"   ✓ Professional formatting")
        
        return output_file
        
    except Exception as e:
        logger.error(f"❌ Error creating Excel file: {str(e)}")
        raise

if __name__ == "__main__":
    create_quarterly_results_excel()
