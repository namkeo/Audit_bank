# -*- coding: utf-8 -*-
"""
Generate Comprehensive Component Analysis Excel Report
Analyzes existing bank audit reports and creates detailed Excel output
"""

import pandas as pd
import json
import os
import sys

# First, check if openpyxl is installed
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    print("⚠ openpyxl not installed. Installing now...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True


def explain_score(risk_type: str, score: float) -> str:
    """Generate explanation for a risk component score"""
    if score < 20:
        level = "Very Low"
        explanation = f"{risk_type} is well-managed with strong controls and minimal concerns."
    elif score < 40:
        level = "Low"
        explanation = f"{risk_type} is generally well-controlled with minor areas for improvement."
    elif score < 60:
        level = "Moderate"
        explanation = f"{risk_type} shows some concerns that require monitoring and potential action."
    elif score < 80:
        level = "High"
        explanation = f"{risk_type} presents significant concerns requiring immediate attention and remediation."
    else:
        level = "Critical"
        explanation = f"{risk_type} is at critical levels requiring urgent intervention and comprehensive action plan."
    
    return f"{level} ({score:.1f}/100) - {explanation}"


def generate_overall_assessment(bank_id: str, overall_score: float, risk_level: str,
                                credit: float, liquidity: float, anomaly: float) -> str:
    """Generate overall assessment narrative"""
    
    highest_risk = max([('Credit', credit), ('Liquidity', liquidity), ('Anomaly', anomaly)], key=lambda x: x[1])
    lowest_risk = min([('Credit', credit), ('Liquidity', liquidity), ('Anomaly', anomaly)], key=lambda x: x[1])
    
    assessment = f"{bank_id} has an overall risk score of {overall_score:.1f} classified as {risk_level}. "
    
    assessment += f"The primary risk driver is {highest_risk[0]} Risk ({highest_risk[1]:.1f}), "
    assessment += f"while {lowest_risk[0]} Risk ({lowest_risk[1]:.1f}) is the strongest performing area. "
    
    if overall_score < 30:
        assessment += "The bank demonstrates strong risk management across all areas with minimal supervisory concerns."
    elif overall_score < 50:
        assessment += "The bank shows generally sound practices but should address identified weaknesses proactively."
    elif overall_score < 70:
        assessment += "The bank requires enhanced monitoring and should implement recommended improvements promptly."
    else:
        assessment += "The bank faces significant risks requiring immediate management attention and regulatory oversight."
    
    return assessment


def generate_bank_recommendations(bank_id: str, credit: float, liquidity: float, 
                                  anomaly: float, overall: float) -> list:
    """Generate specific recommendations for a bank"""
    recommendations = []
    
    # Credit Risk Recommendations
    if credit >= 70:
        recommendations.append({
            'priority': 'HIGH',
            'area': 'Credit Risk',
            'issue': 'Elevated credit risk levels detected',
            'recommendation': 'Conduct comprehensive loan portfolio review, strengthen underwriting standards, and increase loan loss provisions. Consider restricting new high-risk lending.',
            'timeline': 'Immediate - 30 days'
        })
    elif credit >= 50:
        recommendations.append({
            'priority': 'MEDIUM',
            'area': 'Credit Risk',
            'issue': 'Moderate credit risk concerns',
            'recommendation': 'Review credit policies, enhance monitoring of at-risk borrowers, and improve early warning systems for portfolio quality deterioration.',
            'timeline': '30-60 days'
        })
    elif credit >= 30:
        recommendations.append({
            'priority': 'LOW',
            'area': 'Credit Risk',
            'issue': 'Minor credit risk improvements needed',
            'recommendation': 'Maintain current credit standards and continue regular portfolio monitoring. Consider minor enhancements to risk rating systems.',
            'timeline': '60-90 days'
        })
    else:
        recommendations.append({
            'priority': 'LOW',
            'area': 'Credit Risk',
            'issue': 'Strong credit risk management',
            'recommendation': 'Continue existing practices. Share best practices with other institutions. Consider slight expansion within risk appetite.',
            'timeline': 'Ongoing'
        })
    
    # Liquidity Risk Recommendations
    if liquidity >= 70:
        recommendations.append({
            'priority': 'HIGH',
            'area': 'Liquidity Risk',
            'issue': 'Critical liquidity concerns',
            'recommendation': 'Immediately improve liquidity position by increasing liquid assets, reducing loan-to-deposit ratio, and establishing contingency funding plans. Consider asset sales if necessary.',
            'timeline': 'Immediate'
        })
    elif liquidity >= 50:
        recommendations.append({
            'priority': 'MEDIUM',
            'area': 'Liquidity Risk',
            'issue': 'Moderate liquidity pressures',
            'recommendation': 'Strengthen liquidity buffers, diversify funding sources, and improve cash flow forecasting. Review and update contingency funding plan.',
            'timeline': '30-60 days'
        })
    elif liquidity >= 30:
        recommendations.append({
            'priority': 'LOW',
            'area': 'Liquidity Risk',
            'issue': 'Adequate liquidity with room for improvement',
            'recommendation': 'Optimize liquid asset allocation and enhance liquidity stress testing. Consider term structure of assets and liabilities.',
            'timeline': '60-90 days'
        })
    else:
        recommendations.append({
            'priority': 'LOW',
            'area': 'Liquidity Risk',
            'issue': 'Strong liquidity position',
            'recommendation': 'Maintain robust liquidity buffers while optimizing returns. Continue regular stress testing and scenario analysis.',
            'timeline': 'Ongoing'
        })
    
    # Anomaly Detection Recommendations
    if anomaly >= 10:
        recommendations.append({
            'priority': 'HIGH',
            'area': 'Anomaly Detection',
            'issue': 'Significant anomalies detected',
            'recommendation': 'Conduct thorough investigation of all flagged transactions and activities. Enhance internal controls and fraud detection systems. Consider external audit.',
            'timeline': '7-14 days'
        })
    elif anomaly >= 5:
        recommendations.append({
            'priority': 'MEDIUM',
            'area': 'Anomaly Detection',
            'issue': 'Moderate unusual patterns identified',
            'recommendation': 'Review flagged activities, strengthen transaction monitoring, and improve exception reporting processes.',
            'timeline': '30 days'
        })
    elif anomaly > 0:
        recommendations.append({
            'priority': 'LOW',
            'area': 'Anomaly Detection',
            'issue': 'Minor anomalies detected',
            'recommendation': 'Document and review identified anomalies. Enhance monitoring thresholds and validation procedures.',
            'timeline': '60 days'
        })
    else:
        recommendations.append({
            'priority': 'LOW',
            'area': 'Anomaly Detection',
            'issue': 'No significant anomalies',
            'recommendation': 'Continue robust monitoring and detection systems. Regularly update detection algorithms and thresholds.',
            'timeline': 'Ongoing'
        })
    
    # Overall Risk Recommendation
    if overall >= 70:
        recommendations.append({
            'priority': 'CRITICAL',
            'area': 'Overall Risk Management',
            'issue': 'High composite risk requiring urgent attention',
            'recommendation': 'Implement comprehensive risk reduction program. Increase Board and senior management oversight. Consider restrictions on growth and dividends until risk profile improves.',
            'timeline': 'Immediate'
        })
    elif overall >= 50:
        recommendations.append({
            'priority': 'HIGH',
            'area': 'Overall Risk Management',
            'issue': 'Elevated overall risk levels',
            'recommendation': 'Develop and implement risk mitigation action plan with clear timelines and accountability. Enhance risk reporting to Board.',
            'timeline': '30 days'
        })
    
    return recommendations


def _attach_quantile_classification(df_components: pd.DataFrame, summary_path: str = "outputs/all_banks_summary.csv") -> pd.DataFrame:
    """Attach quantile classification using the precomputed summary when available."""
    quantile_col = "Quantile Classification"
    quantile_map = {}

    if os.path.exists(summary_path):
        try:
            summary_df = pd.read_csv(summary_path)
            id_col = 'bank_id' if 'bank_id' in summary_df.columns else 'Bank ID'
            qc_col = 'quantile_classification' if 'quantile_classification' in summary_df.columns else quantile_col
            if id_col in summary_df.columns and qc_col in summary_df.columns:
                summary_df['Bank ID'] = summary_df[id_col].astype(str).str.upper()
                quantile_map = dict(zip(summary_df['Bank ID'], summary_df[qc_col]))
                print(f"[+] Loaded quantile classifications from {summary_path}")
        except Exception as exc:
            print(f"⚠ Could not read {summary_path}: {exc}")

    if quantile_col not in df_components:
        df_components[quantile_col] = pd.Series(index=df_components.index, dtype=object)

    if quantile_map:
        df_components[quantile_col] = df_components['Bank ID'].map(quantile_map)

    thresholds = df_components['Overall Score'].quantile([0.25, 0.5, 0.75])

    def classify(score: float) -> str:
        if pd.isna(score):
            return "UNKNOWN"
        if score <= thresholds.loc[0.25]:
            return "MINIMAL"
        if score <= thresholds.loc[0.50]:
            return "LOW"
        if score <= thresholds.loc[0.75]:
            return "MEDIUM"
        return "HIGH"

    df_components[quantile_col] = df_components[quantile_col].fillna(
        df_components['Overall Score'].apply(classify)
    )

    ordered_cols = ['Bank ID', 'Overall Score', 'Risk Level', quantile_col]
    remaining_cols = [col for col in df_components.columns if col not in ordered_cols]
    return df_components[ordered_cols + remaining_cols]


def format_component_sheet(worksheet):
    """Format the Component Scores sheet"""
    # Header formatting
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Column widths
    worksheet.column_dimensions['A'].width = 12   # Bank ID
    worksheet.column_dimensions['B'].width = 15   # Overall Score
    worksheet.column_dimensions['C'].width = 15   # Risk Level
    worksheet.column_dimensions['D'].width = 24   # Quantile Classification
    worksheet.column_dimensions['E'].width = 18   # Credit Risk Score
    worksheet.column_dimensions['F'].width = 20   # Liquidity Risk Score
    worksheet.column_dimensions['G'].width = 15   # Anomaly Score
    worksheet.column_dimensions['H'].width = 20   # Highest Risk Area
    
    # Freeze header row
    worksheet.freeze_panes = 'A2'


def format_analysis_sheet(worksheet):
    """Format the Risk Analysis sheet"""
    # Header formatting
    header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Column widths
    worksheet.column_dimensions['A'].width = 12
    worksheet.column_dimensions['B'].width = 50
    worksheet.column_dimensions['C'].width = 50
    worksheet.column_dimensions['D'].width = 50
    worksheet.column_dimensions['E'].width = 60
    
    # Text wrapping for all cells
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    worksheet.freeze_panes = 'A2'


def format_recommendations_sheet(worksheet):
    """Format the Recommendations sheet"""
    # Header formatting
    header_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Column widths
    worksheet.column_dimensions['A'].width = 12
    worksheet.column_dimensions['B'].width = 12
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 40
    worksheet.column_dimensions['E'].width = 60
    worksheet.column_dimensions['F'].width = 18
    
    # Text wrapping
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Color-code priorities
    for row in worksheet.iter_rows(min_row=2):
        priority_cell = row[1]  # Column B
        if priority_cell.value == 'CRITICAL':
            for cell in row:
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        elif priority_cell.value == 'HIGH':
            for cell in row:
                cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        elif priority_cell.value == 'MEDIUM':
            for cell in row:
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    
    worksheet.freeze_panes = 'A2'


def main():
    """Main function to generate Excel analysis"""
    print("="*80)
    print("GENERATING COMPREHENSIVE COMPONENT ANALYSIS")
    print("="*80)
    
    output_file = "outputs/banks_components_analysis.xlsx"
    # Ensure only one Excel output remains in the folder
    for fname in os.listdir("outputs"):
        if fname.lower().endswith(".xlsx") and fname != os.path.basename(output_file):
            try:
                os.remove(os.path.join("outputs", fname))
                print(f"Removed old Excel file: outputs/{fname}")
            except OSError as exc:
                print(f"⚠ Could not remove outputs/{fname}: {exc}")
    
    # Create outputs directory if needed
    os.makedirs("outputs", exist_ok=True)
    
    # Load all bank reports
    bank_reports = {}
    for file in os.listdir("outputs"):
        if file.endswith("_bank_report.json"):
            bank_id = file.replace("_bank_report.json", "").upper()
            with open(f"outputs/{file}", 'r', encoding='utf-8') as f:
                bank_reports[bank_id] = json.load(f)
    
    print(f"[+] Loaded {len(bank_reports)} bank reports")
    
    # Prepare data structures
    component_data = []
    explanations_data = []
    recommendations_data = []
    
    for bank_id, report in sorted(bank_reports.items()):
        overall = report.get('overall_risk_score', {})
        components = overall.get('component_scores', {})
        risk_level = overall.get('risk_level', 'UNKNOWN')
        overall_score = overall.get('overall_score', 0)
        
        # Component scores
        credit_score = components.get('credit', 0)
        liquidity_score = components.get('liquidity', 0)
        anomaly_score = components.get('anomaly', 0)
        
        component_data.append({
            'Bank ID': bank_id,
            'Overall Score': round(overall_score, 2),
            'Risk Level': risk_level,
            'Credit Risk Score': round(credit_score, 2),
            'Liquidity Risk Score': round(liquidity_score, 2),
            'Anomaly Score': round(anomaly_score, 2),
            'Highest Risk Area': max([
                ('Credit', credit_score),
                ('Liquidity', liquidity_score),
                ('Anomaly', anomaly_score)
            ], key=lambda x: x[1])[0]
        })
        
        # Generate explanations
        credit_explanation = explain_score('Credit Risk', credit_score)
        liquidity_explanation = explain_score('Liquidity Risk', liquidity_score)
        anomaly_explanation = explain_score('Anomaly Detection', anomaly_score)
        
        explanations_data.append({
            'Bank ID': bank_id,
            'Credit Risk Analysis': credit_explanation,
            'Liquidity Risk Analysis': liquidity_explanation,
            'Anomaly Analysis': anomaly_explanation,
            'Overall Assessment': generate_overall_assessment(
                bank_id, overall_score, risk_level, credit_score, liquidity_score, anomaly_score
            )
        })
        
        # Generate recommendations
        bank_recommendations = generate_bank_recommendations(
            bank_id, credit_score, liquidity_score, anomaly_score, overall_score
        )
        
        for rec in bank_recommendations:
            recommendations_data.append({
                'Bank ID': bank_id,
                'Priority': rec['priority'],
                'Risk Area': rec['area'],
                'Issue': rec['issue'],
                'Recommendation': rec['recommendation'],
                'Timeline': rec['timeline']
            })
    
    # Create Excel file with multiple sheets
    print(f"\n[+] Creating Excel file: {output_file}")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Sheet 1: Component Scores Summary
        df_components = pd.DataFrame(component_data)
        df_components = df_components.sort_values('Overall Score', ascending=False)
        df_components = _attach_quantile_classification(df_components)
        df_components.to_excel(writer, sheet_name='Component Scores', index=False)
        
        # Sheet 2: Detailed Explanations
        df_explanations = pd.DataFrame(explanations_data)
        df_explanations.to_excel(writer, sheet_name='Risk Analysis', index=False)
        
        # Sheet 3: Recommendations
        df_recommendations = pd.DataFrame(recommendations_data)
        df_recommendations = df_recommendations.sort_values(['Priority', 'Bank ID'])
        df_recommendations.to_excel(writer, sheet_name='Recommendations', index=False)
        
        # Format the sheets
        workbook = writer.book
        format_component_sheet(workbook['Component Scores'])
        format_analysis_sheet(workbook['Risk Analysis'])
        format_recommendations_sheet(workbook['Recommendations'])
    
    print(f"[+] Successfully created: {output_file}")
    print(f"\n📊 Excel file contains 3 sheets:")
    print(f"  1. Component Scores - Rankings and scores for all {len(bank_reports)} banks")
    print(f"  2. Risk Analysis - Detailed explanations for each risk component")
    print(f"  3. Recommendations - {len(recommendations_data)} actionable recommendations by priority")
    
    # Clean up old CSV files that were replaced by Excel
    for fname in ['banks_component_scores.csv', 'banks_risk_analysis.csv', 'banks_recommendations.csv']:
        fpath = os.path.join('outputs', fname)
        if os.path.exists(fpath):
            try:
                os.remove(fpath)
                print(f"Removed old CSV file: {fname}")
            except OSError as exc:
                print(f"⚠ Could not remove {fname}: {exc}")
    
    # Print summary
    print(f"\n" + "="*80)
    print("SUMMARY OF FINDINGS")
    print("="*80)
    
    df_comp = pd.DataFrame(component_data).sort_values('Overall Score', ascending=False)
    print("\nTop 5 Banks by Risk (Highest Risk):")
    for idx, row in enumerate(df_comp.head().itertuples(), 1):
        print(f"  {idx}. {row._1}: {row._2:.2f} ({row._3}) - Highest concern: {row._7}")
    
    print("\nLowest Risk Banks:")
    for idx, row in enumerate(df_comp.tail().itertuples(), 1):
        print(f"  {idx}. {row._1}: {row._2:.2f} ({row._3})")
    
    # Priority recommendations count
    priority_counts = pd.DataFrame(recommendations_data)['Priority'].value_counts()
    print(f"\nRecommendations by Priority:")
    for priority in ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW']:
        count = priority_counts.get(priority, 0)
        if count > 0:
            print(f"  {priority}: {count} recommendations")
    
    print("\n" + "="*80)
    print("✅ Analysis complete! Open the Excel file to view detailed results.")
    print("="*80)


if __name__ == "__main__":
    main()
