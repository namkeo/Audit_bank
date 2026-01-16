# -*- coding: utf-8 -*-
"""
Example Usage of Modular Bank Audit System
Demonstrates how to use the refactored components
"""

import pandas as pd
import numpy as np
import sys
import os
import logging
import warnings
import importlib.util
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment

# Suppress sklearn robust covariance warnings
warnings.filterwarnings('ignore', category=RuntimeWarning, message='Determinant has increased')
warnings.filterwarnings('ignore', category=UserWarning, message='The covariance matrix associated to your dataset is not full rank')

# Import quarterly scores generation
spec_quarterly = importlib.util.spec_from_file_location("quarterly_generator", "generate_quarterly_excel.py")
quarterly_generator = importlib.util.module_from_spec(spec_quarterly)
spec_quarterly.loader.exec_module(quarterly_generator)

spec_quarterly_scores = importlib.util.spec_from_file_location("quarterly_scores_gen", "generate_quarterly_scores.py")
quarterly_scores_gen = importlib.util.module_from_spec(spec_quarterly_scores)
spec_quarterly_scores.loader.exec_module(quarterly_scores_gen)

# Setup logging - import from 6_logging_config.py
spec_logging = importlib.util.spec_from_file_location("logging_config", "6_logging_config.py")
logging_module = importlib.util.module_from_spec(spec_logging)
spec_logging.loader.exec_module(logging_module)
AuditLogger = logging_module.AuditLogger

AuditLogger.setup_logging(log_level="INFO", console_output=True)
logger = AuditLogger.get_logger(__name__)

# Import the BankAuditSystem 

import importlib.util
spec = importlib.util.spec_from_file_location("bank_audit_system", "5_bank_audit_system.py")
bank_audit_module = importlib.util.module_from_spec(spec)
spec.loader.exec_module(bank_audit_module)
BankAuditSystem = bank_audit_module.BankAuditSystem

# Prepare dataset: use enhanced dataset with concentration and OBS indicators
try:
    # Check if we have the latest enriched dataset with new indicators
    if Path("data/time_series_dataset_enriched_v2.csv").exists():
        DATASET_CSV = "data/time_series_dataset_enriched_v2.csv"
        logger.info(f"Using enhanced dataset with concentration/OBS indicators: {DATASET_CSV}")
    elif Path("time_series_dataset_enriched_v2.csv").exists():
        DATASET_CSV = "time_series_dataset_enriched_v2.csv"
        logger.info(f"Using enhanced dataset (root): {DATASET_CSV}")
    else:
        # Fallback to original enriched dataset
        from utils.dataset_enrichment import enrich_time_series_dataset
        DATASET_CSV, _ = enrich_time_series_dataset()
        logger.info(f"Using enriched dataset: {DATASET_CSV}")
except Exception as e:
    DATASET_CSV = 'data/time_series_dataset.csv'
    logger.warning(f"Could not load enriched dataset, using base dataset: {DATASET_CSV}. Error: {e}")


# ============================================================================
# EXAMPLE 1: Simple Audit for a Single Bank
# ============================================================================

def run_simple_audit_example():
    """
    Example of running a simple audit for one bank
    """
    logger.info("="*80)
    logger.info("EXAMPLE 1: Simple Audit for Single Bank")
    logger.info("="*80)
    
    # Load your data
    df = pd.read_csv(DATASET_CSV)
    logger.debug(f"Loaded {len(df)} records from {DATASET_CSV}")
    
    # Initialize the audit system
    audit_system = BankAuditSystem(
        bank_name="VCB Bank",
        audit_period="2024-Q1"
    )
    
    # Run complete audit
    report = audit_system.run_complete_audit(
        df=df,
        bank_id="VCB",  # Use a valid bank_id from the dataset
        all_banks_data=df  # Use full dataset for training
    )
    
    # Print summary
    audit_system.print_summary(report)
    
    # Create dashboard
    audit_system.create_dashboard(report, save_path="outputs/vcb_bank_dashboard.png")
    
    # Export results
    audit_system.export_results("outputs/vcb_bank_report.json", format='json')
    # audit_system.export_results("outputs/vcb_bank_report.xlsx", format='excel')  # Skip if file is open
    
    return report


# ============================================================================
# EXAMPLE 2: Detailed Step-by-Step Audit
# ============================================================================

def run_detailed_audit_example():
    """
    Example showing detailed control over each step
    """
    logger.info("EXAMPLE 2: Detailed Step-by-Step Audit")
    
    # Load data
    df = pd.read_csv(DATASET_CSV)
    
    # Initialize system
    audit_system = BankAuditSystem(
        bank_name="BIDV Bank",
        audit_period="2024-Q2"
    )
    
    # Step 1: Prepare data
    logger.info("Step 1: Preparing data...")
    prepared_data = audit_system.load_and_prepare_data(df, bank_id="BIDV")
    
    logger.info(f"  - Loaded {len(prepared_data['ratios'])} periods")
    logger.info(f"  - Calculated {len(prepared_data['features'].columns)} features")
    
    # Step 2: Train models
    logger.info("Step 2: Training models...")
    training_results = audit_system.train_models(df)
    
    for model_type, results in training_results.items():
        logger.debug(f"  - {model_type}: {results}")
    
    # Step 3: Assess risks
    logger.info("Step 3: Assessing risks...")
    risk_assessments = audit_system.assess_all_risks(prepared_data)
    
    logger.info(f"  - Credit Risk Score: {risk_assessments['credit_risk'].get('credit_risk_score', 0):.2f}")
    logger.info(f"  - Liquidity Risk Score: {risk_assessments['liquidity_risk'].get('liquidity_risk_score', 0):.2f}")
    logger.info(f"  - Anomalies Detected: {risk_assessments['anomaly_detection'].get('anomalies_detected', 0)}")
    
    # Step 4: Apply expert rules
    logger.info("Step 4: Applying expert rules...")
    violations = audit_system.apply_expert_rules(prepared_data)
    
    logger.info(f"  - Found {len(violations)} rule violations")
    for violation in violations:
        logger.warning(f"    RULE VIOLATION: {violation['message']}")
    
    # Step 5: Identify high risk periods
    logger.info("Step 5: Identifying high risk periods...")
    high_risk_periods = audit_system.identify_high_risk_periods(prepared_data)
    
    logger.info(f"  - Found {len(high_risk_periods)} high risk periods")
    
    # Step 6: Generate report
    logger.info("Step 6: Generating comprehensive report...")
    report = audit_system.generate_comprehensive_report()
    
    # Display results
    audit_system.print_summary(report)
    
    return report


# ============================================================================
# EXAMPLE 3: Using Individual Components
# ============================================================================

def run_component_level_example():
    """
    Example using individual components directly
    """
    logger.info("="*80)
    logger.info("EXAMPLE 3: Using Individual Components")
    logger.info("="*80)
    
    import importlib
    data_prep_module = importlib.import_module('1_data_preparation')
    credit_risk_module = importlib.import_module('2_model_credit_risk')
    liquidity_risk_module = importlib.import_module('2_model_liquidity_risk')
    
    DataPreparation = data_prep_module.DataPreparation
    CreditRiskModel = credit_risk_module.CreditRiskModel
    LiquidityRiskModel = liquidity_risk_module.LiquidityRiskModel
    
    # Load data
    df = pd.read_csv(DATASET_CSV)
    logger.debug(f"Loaded {len(df)} records")
    
    # Use Data Preparation component
    logger.info("Using Data Preparation Component...")
    data_prep = DataPreparation()
    time_series = data_prep.load_time_series_data(df, bank_id="BIDV")
    ratios = data_prep.calculate_time_series_ratios()
    logger.info(f"  - Calculated ratios for {len(ratios)} periods")
    
    # Use Credit Risk component
    logger.info("Using Credit Risk Component...")
    credit_model = CreditRiskModel()
    features = data_prep.prepare_time_series_features()
    
    # Train on all banks data
    credit_model.train_models(features)
    
    # Predict risk
    credit_results = credit_model.predict_risk(features)
    logger.info(f"  - Credit Risk Score: {credit_results['credit_risk_score']:.2f}")
    logger.info(f"  - Risk Level: {credit_results['risk_level']}")
    
    # Assess indicators
    indicators = credit_model.assess_risk_indicators(features)
    logger.info(f"  - Assessed {len(indicators)} credit indicators")
    
    # Use Liquidity Risk component
    logger.info("Using Liquidity Risk Component...")
    liquidity_model = LiquidityRiskModel()
    liquidity_model.train_models(features)
    
    liquidity_results = liquidity_model.predict_risk(features)
    logger.info(f"  - Liquidity Risk Score: {liquidity_results['liquidity_risk_score']:.2f}")
    
    # Run stress tests
    stress_results = liquidity_model.run_liquidity_stress_test(features)
    logger.info(f"  - Stress test scenarios: {len(stress_results)}")
    for scenario, result in stress_results.items():
        status = 'PASS' if result['passes_stress_test'] else 'FAIL'
        log_level = logger.info if result['passes_stress_test'] else logger.warning
        log_level(f"    • {scenario}: {status}")


# ============================================================================
# EXAMPLE 4: Multiple Banks Analysis
# ============================================================================

def run_multiple_banks_example():
    """
    Example analyzing multiple banks
    """
    logger.info("EXAMPLE 4: Multiple Banks Analysis")
    
    # Load data
    df = pd.read_csv(DATASET_CSV)
    
    # Get unique bank IDs
    bank_ids = df['bank_id'].unique()
    
    logger.info(f"Analyzing {len(bank_ids)} banks...")
    
    all_reports = {}
    
    for bank_id in bank_ids[:3]:  # Limit to first 3 for demo
        logger.info(f"Analyzing {bank_id}...")
        
        # Create audit system for this bank
        audit_system = BankAuditSystem(
            bank_name=bank_id,
            audit_period="2024"
        )
        
        # Run audit
        report = audit_system.run_complete_audit(
            df=df[df['bank_id'] == bank_id],
            bank_id=bank_id,
            all_banks_data=df
        )
        
        all_reports[bank_id] = report
        
        # Print brief summary
        overall_score = report.get('overall_risk_score', {}).get('overall_score', 0)
        risk_level = report.get('overall_risk_score', {}).get('risk_level', 'UNKNOWN')
        logger.info(f"  Overall Risk: {overall_score:.2f} ({risk_level})")
    
    # Compare banks
    logger.info("COMPARISON SUMMARY")
    
    for bank_id, report in all_reports.items():
        overall = report.get('overall_risk_score', {})
        logger.info(f"{bank_id:20s}: Score={overall.get('overall_score', 0):6.2f}, Level={overall.get('risk_level', 'UNKNOWN')}")
    
    return all_reports


# ============================================================================
# EXAMPLE: Audit ALL Banks and Save to Outputs
# ============================================================================

def audit_all_banks_to_outputs():
    """
    Audit all banks in the dataset and save results to outputs folder
    """
    logger.info("="*80)
    logger.info("AUDITING ALL BANKS - SAVING TO OUTPUTS FOLDER")
    logger.info("="*80)
    
    # Create outputs directory if it doesn't exist
    os.makedirs("outputs", exist_ok=True)
    
    # Load data
    df = pd.read_csv(DATASET_CSV)
    
    # Get all unique bank IDs
    bank_ids = df['bank_id'].unique()
    
    logger.info(f"Found {len(bank_ids)} banks to audit: {', '.join(bank_ids)}")
    logger.info("")
    
    all_reports = {}
    summary_results = []
    
    for idx, bank_id in enumerate(bank_ids, 1):
        logger.info(f"[{idx}/{len(bank_ids)}] Processing {bank_id}...")
        
        try:
            # Create audit system for this bank
            audit_system = BankAuditSystem(
                bank_name=f"{bank_id} Bank",
                audit_period="2024"
            )
            
            # Run complete audit
            report = audit_system.run_complete_audit(
                df=df,
                bank_id=bank_id,
                all_banks_data=df
            )
            
            all_reports[bank_id] = report
            
            # Get summary metrics
            overall_score = report.get('overall_risk_score', {}).get('overall_score', 0)
            risk_level = report.get('overall_risk_score', {}).get('risk_level', 'UNKNOWN')
            
            summary_results.append({
                'bank_id': bank_id,
                'overall_score': overall_score,
                'risk_level': risk_level
            })
            
            logger.info(f"  ✓ Overall Risk: {overall_score:.2f} ({risk_level})")
            
            # Save outputs for this bank
            output_json = f"outputs/{bank_id.lower()}_bank_report.json"
            output_dashboard = f"outputs/{bank_id.lower()}_bank_dashboard.png"
            
            # Export JSON report
            audit_system.export_results(output_json, format='json')
            logger.info(f"  ✓ Saved JSON: {output_json}")
            
            # Create and save dashboard
            audit_system.create_dashboard(report, save_path=output_dashboard)
            logger.info(f"  ✓ Saved Dashboard: {output_dashboard}")
            
            logger.info("")
            
        except Exception as e:
            logger.error(f"  ✗ Error processing {bank_id}: {str(e)}")
            logger.info("")
            continue
    
    # Calculate quantiles from all collected scores and reclassify
    logger.info("")
    logger.info("="*80)
    logger.info("COMPUTING QUANTILE-BASED RISK CLASSIFICATION")
    logger.info("="*80)
    
    # We need to access a shared ReportingAnalysis instance to compute quantiles
    # For now, we'll compute them separately and create a reference table
    import numpy as np
    all_scores = [result['overall_score'] for result in summary_results]
    q25 = np.quantile(all_scores, 0.25)
    q50 = np.quantile(all_scores, 0.50)
    q75 = np.quantile(all_scores, 0.75)
    
    logger.info(f"\nQuantile Thresholds (from {len(all_scores)} banks):")
    logger.info(f"  Q1 (25th percentile): {q25:.2f}")
    logger.info(f"  Q2 (50th percentile): {q50:.2f}")
    logger.info(f"  Q3 (75th percentile): {q75:.2f}")
    
    # Classify by quantiles
    def classify_by_quantile(score):
        if score <= q25:
            return "MINIMAL"
        elif score <= q50:
            return "LOW"
        elif score <= q75:
            return "MEDIUM"
        else:
            return "HIGH"
    
    for result in summary_results:
        result['quantile_classification'] = classify_by_quantile(result['overall_score'])
    
    # Print final summary
    logger.info("="*80)
    logger.info("AUDIT SUMMARY - ALL BANKS (Quantile-Based Classification)")
    logger.info("="*80)
    
    summary_df = pd.DataFrame(summary_results)
    summary_df = summary_df.sort_values('overall_score', ascending=False)
    
    logger.info("\nRanking by Risk Score (Highest to Lowest):")
    logger.info("Bank  | Score  | Fixed Method | Quantile Method")
    logger.info("-" * 55)
    for idx, row in enumerate(summary_df.itertuples(), 1):
        logger.info(f"{row.bank_id:6s} | {row.overall_score:6.2f} | {row.risk_level:12s} | {row.quantile_classification}")
    
    # Save consolidated summary
    summary_csv = "outputs/all_banks_summary.csv"
    summary_df.to_csv(summary_csv, index=False)
    logger.info(f"\n✓ Consolidated summary saved to: {summary_csv}")
    
    logger.info("="*80)
    logger.info(f"COMPLETED: Audited {len(all_reports)}/{len(bank_ids)} banks")
    logger.info(f"Results saved to outputs/ folder")
    logger.info("="*80)
    
    return all_reports


# ============================================================================
# COMPREHENSIVE COMPONENT ANALYSIS TO EXCEL
# ============================================================================

def export_component_analysis_to_excel(output_file: str = "outputs/banks_components_analysis.xlsx"):
    """
    Export detailed component scores, explanations, and recommendations to Excel
    
    Args:
        output_file: Path to output Excel file
    """
    import json
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    logger.info("="*80)
    logger.info("GENERATING COMPREHENSIVE COMPONENT ANALYSIS")
    logger.info("="*80)
    
    # Create outputs directory if needed
    os.makedirs("outputs", exist_ok=True)

    # Ensure only the consolidated Excel remains in outputs
    for fname in os.listdir("outputs"):
        if fname.lower().endswith(".xlsx") and fname != os.path.basename(output_file):
            try:
                os.remove(os.path.join("outputs", fname))
                logger.info(f"Removed old Excel file: outputs/{fname}")
            except OSError as exc:
                logger.warning(f"Could not remove outputs/{fname}: {exc}")
    
    # Load all bank reports
    bank_reports = {}
    for file in os.listdir("outputs"):
        if file.endswith("_bank_report.json"):
            bank_id = file.replace("_bank_report.json", "").upper()
            with open(f"outputs/{file}", 'r', encoding='utf-8') as f:
                bank_reports[bank_id] = json.load(f)
    
    logger.info(f"Loaded {len(bank_reports)} bank reports")
    
    # Prepare data structures
    component_data = []
    explanations_data = []
    recommendations_data = []
    
    for bank_id, report in sorted(bank_reports.items()):
        overall = report.get('overall_risk_score', {})
        components = overall.get('component_scores', {})
        risk_level = overall.get('risk_level', 'UNKNOWN')
        overall_score = overall.get('overall_score', 0)
        
        # Component scores
        credit_score = components.get('credit', 0)
        liquidity_score = components.get('liquidity', 0)
        anomaly_score = components.get('anomaly', 0)
        
        component_data.append({
            'Bank ID': bank_id,
            'Overall Score': round(overall_score, 2),
            'Risk Level': risk_level,
            'Credit Risk Score': round(credit_score, 2),
            'Liquidity Risk Score': round(liquidity_score, 2),
            'Anomaly Score': round(anomaly_score, 2),
            'Highest Risk Area': max([
                ('Credit', credit_score),
                ('Liquidity', liquidity_score),
                ('Anomaly', anomaly_score)
            ], key=lambda x: x[1])[0]
        })
        
        # Generate explanations
        credit_explanation = _explain_score('Credit Risk', credit_score)
        liquidity_explanation = _explain_score('Liquidity Risk', liquidity_score)
        anomaly_explanation = _explain_score('Anomaly Detection', anomaly_score)
        
        explanations_data.append({
            'Bank ID': bank_id,
            'Credit Risk Analysis': credit_explanation,
            'Liquidity Risk Analysis': liquidity_explanation,
            'Anomaly Analysis': anomaly_explanation,
            'Overall Assessment': _generate_overall_assessment(
                bank_id, overall_score, risk_level, credit_score, liquidity_score, anomaly_score
            )
        })
        
        # Generate recommendations
        bank_recommendations = _generate_bank_recommendations(
            bank_id, credit_score, liquidity_score, anomaly_score, overall_score
        )
        
        for rec in bank_recommendations:
            recommendations_data.append({
                'Bank ID': bank_id,
                'Priority': rec['priority'],
                'Risk Area': rec['area'],
                'Issue': rec['issue'],
                'Recommendation': rec['recommendation'],
                'Timeline': rec['timeline']
            })
    
    # Create Excel file with multiple sheets
    logger.info(f"Creating Excel file: {output_file}")
    
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Sheet 1: Component Scores Summary
            df_components = pd.DataFrame(component_data)
            df_components = df_components.sort_values('Overall Score', ascending=False)
            df_components = _attach_quantile_classification(df_components)
            df_components.to_excel(writer, sheet_name='Component Scores', index=False)
            
            # Sheet 2: Detailed Explanations
            df_explanations = pd.DataFrame(explanations_data)
            df_explanations.to_excel(writer, sheet_name='Risk Analysis', index=False)
            
            # Sheet 3: Recommendations
            df_recommendations = pd.DataFrame(recommendations_data)
            df_recommendations = df_recommendations.sort_values(['Priority', 'Bank ID'])
            df_recommendations.to_excel(writer, sheet_name='Recommendations', index=False)
            
            # Format the sheets
            workbook = writer.book
            _format_component_sheet(workbook['Component Scores'])
            _format_analysis_sheet(workbook['Risk Analysis'])
            _format_recommendations_sheet(workbook['Recommendations'])
        
        logger.info(f"✓ Successfully created: {output_file}")
        logger.info(f"  - Component Scores: Rankings and scores for all banks")
        logger.info(f"  - Risk Analysis: Detailed explanations for each bank")
        logger.info(f"  - Recommendations: Actionable recommendations by priority")
        
    except Exception as e:
        logger.error(f"Error creating Excel file: {e}")
        raise
    
    # Clean up old CSV files that were replaced by Excel
    for fname in ['banks_component_scores.csv', 'banks_risk_analysis.csv', 'banks_recommendations.csv']:
        fpath = os.path.join('outputs', fname)
        if os.path.exists(fpath):
            try:
                os.remove(fpath)
                logger.info(f"Removed old CSV file: {fname}")
            except OSError as exc:
                logger.warning(f"Could not remove {fname}: {exc}")
    
    logger.info("="*80)
    return component_data, explanations_data, recommendations_data


def _explain_score(risk_type: str, score: float) -> str:
    """Generate explanation for a risk component score"""
    if score < 20:
        level = "Very Low"
        explanation = f"{risk_type} is well-managed with strong controls and minimal concerns."
    elif score < 40:
        level = "Low"
        explanation = f"{risk_type} is generally well-controlled with minor areas for improvement."
    elif score < 60:
        level = "Moderate"
        explanation = f"{risk_type} shows some concerns that require monitoring and potential action."
    elif score < 80:
        level = "High"
        explanation = f"{risk_type} presents significant concerns requiring immediate attention and remediation."
    else:
        level = "Critical"
        explanation = f"{risk_type} is at critical levels requiring urgent intervention and comprehensive action plan."
    
    return f"{level} ({score:.1f}/100) - {explanation}"


def _generate_overall_assessment(bank_id: str, overall_score: float, risk_level: str,
                                 credit: float, liquidity: float, anomaly: float) -> str:
    """Generate overall assessment narrative"""
    
    highest_risk = max([('Credit', credit), ('Liquidity', liquidity), ('Anomaly', anomaly)], key=lambda x: x[1])
    lowest_risk = min([('Credit', credit), ('Liquidity', liquidity), ('Anomaly', anomaly)], key=lambda x: x[1])
    
    assessment = f"{bank_id} has an overall risk score of {overall_score:.1f} classified as {risk_level}. "
    
    assessment += f"The primary risk driver is {highest_risk[0]} Risk ({highest_risk[1]:.1f}), "
    assessment += f"while {lowest_risk[0]} Risk ({lowest_risk[1]:.1f}) is the strongest performing area. "
    
    if overall_score < 30:
        assessment += "The bank demonstrates strong risk management across all areas with minimal supervisory concerns."
    elif overall_score < 50:
        assessment += "The bank shows generally sound practices but should address identified weaknesses proactively."
    elif overall_score < 70:
        assessment += "The bank requires enhanced monitoring and should implement recommended improvements promptly."
    else:
        assessment += "The bank faces significant risks requiring immediate management attention and regulatory oversight."
    
    return assessment


def _generate_bank_recommendations(bank_id: str, credit: float, liquidity: float, 
                                  anomaly: float, overall: float) -> list:
    """Generate specific recommendations for a bank"""
    recommendations = []
    
    # Credit Risk Recommendations
    if credit >= 70:
        recommendations.append({
            'priority': 'HIGH',
            'area': 'Credit Risk',
            'issue': 'Elevated credit risk levels detected',
            'recommendation': 'Conduct comprehensive loan portfolio review, strengthen underwriting standards, and increase loan loss provisions. Consider restricting new high-risk lending.',
            'timeline': 'Immediate - 30 days'
        })
    elif credit >= 50:
        recommendations.append({
            'priority': 'MEDIUM',
            'area': 'Credit Risk',
            'issue': 'Moderate credit risk concerns',
            'recommendation': 'Review credit policies, enhance monitoring of at-risk borrowers, and improve early warning systems for portfolio quality deterioration.',
            'timeline': '30-60 days'
        })
    elif credit >= 30:
        recommendations.append({
            'priority': 'LOW',
            'area': 'Credit Risk',
            'issue': 'Minor credit risk improvements needed',
            'recommendation': 'Maintain current credit standards and continue regular portfolio monitoring. Consider minor enhancements to risk rating systems.',
            'timeline': '60-90 days'
        })
    else:
        recommendations.append({
            'priority': 'LOW',
            'area': 'Credit Risk',
            'issue': 'Strong credit risk management',
            'recommendation': 'Continue existing practices. Share best practices with other institutions. Consider slight expansion within risk appetite.',
            'timeline': 'Ongoing'
        })
    
    # Liquidity Risk Recommendations
    if liquidity >= 70:
        recommendations.append({
            'priority': 'HIGH',
            'area': 'Liquidity Risk',
            'issue': 'Critical liquidity concerns',
            'recommendation': 'Immediately improve liquidity position by increasing liquid assets, reducing loan-to-deposit ratio, and establishing contingency funding plans. Consider asset sales if necessary.',
            'timeline': 'Immediate'
        })
    elif liquidity >= 50:
        recommendations.append({
            'priority': 'MEDIUM',
            'area': 'Liquidity Risk',
            'issue': 'Moderate liquidity pressures',
            'recommendation': 'Strengthen liquidity buffers, diversify funding sources, and improve cash flow forecasting. Review and update contingency funding plan.',
            'timeline': '30-60 days'
        })
    elif liquidity >= 30:
        recommendations.append({
            'priority': 'LOW',
            'area': 'Liquidity Risk',
            'issue': 'Adequate liquidity with room for improvement',
            'recommendation': 'Optimize liquid asset allocation and enhance liquidity stress testing. Consider term structure of assets and liabilities.',
            'timeline': '60-90 days'
        })
    else:
        recommendations.append({
            'priority': 'LOW',
            'area': 'Liquidity Risk',
            'issue': 'Strong liquidity position',
            'recommendation': 'Maintain robust liquidity buffers while optimizing returns. Continue regular stress testing and scenario analysis.',
            'timeline': 'Ongoing'
        })
    
    # Anomaly Detection Recommendations
    if anomaly >= 10:
        recommendations.append({
            'priority': 'HIGH',
            'area': 'Anomaly Detection',
            'issue': 'Significant anomalies detected',
            'recommendation': 'Conduct thorough investigation of all flagged transactions and activities. Enhance internal controls and fraud detection systems. Consider external audit.',
            'timeline': '7-14 days'
        })
    elif anomaly >= 5:
        recommendations.append({
            'priority': 'MEDIUM',
            'area': 'Anomaly Detection',
            'issue': 'Moderate unusual patterns identified',
            'recommendation': 'Review flagged activities, strengthen transaction monitoring, and improve exception reporting processes.',
            'timeline': '30 days'
        })
    elif anomaly > 0:
        recommendations.append({
            'priority': 'LOW',
            'area': 'Anomaly Detection',
            'issue': 'Minor anomalies detected',
            'recommendation': 'Document and review identified anomalies. Enhance monitoring thresholds and validation procedures.',
            'timeline': '60 days'
        })
    else:
        recommendations.append({
            'priority': 'LOW',
            'area': 'Anomaly Detection',
            'issue': 'No significant anomalies',
            'recommendation': 'Continue robust monitoring and detection systems. Regularly update detection algorithms and thresholds.',
            'timeline': 'Ongoing'
        })
    
    # Overall Risk Recommendation
    if overall >= 70:
        recommendations.append({
            'priority': 'CRITICAL',
            'area': 'Overall Risk Management',
            'issue': 'High composite risk requiring urgent attention',
            'recommendation': 'Implement comprehensive risk reduction program. Increase Board and senior management oversight. Consider restrictions on growth and dividends until risk profile improves.',
            'timeline': 'Immediate'
        })
    elif overall >= 50:
        recommendations.append({
            'priority': 'HIGH',
            'area': 'Overall Risk Management',
            'issue': 'Elevated overall risk levels',
            'recommendation': 'Develop and implement risk mitigation action plan with clear timelines and accountability. Enhance risk reporting to Board.',
            'timeline': '30 days'
        })
    
    return recommendations


def _attach_quantile_classification(df_components: pd.DataFrame, summary_path: str = "outputs/all_banks_summary.csv") -> pd.DataFrame:
    """Attach quantile classification to the component scores sheet."""
    quantile_col = "Quantile Classification"
    quantile_map = {}

    if os.path.exists(summary_path):
        try:
            summary_df = pd.read_csv(summary_path)
            id_col = 'bank_id' if 'bank_id' in summary_df.columns else 'Bank ID'
            qc_col = 'quantile_classification' if 'quantile_classification' in summary_df.columns else quantile_col
            if id_col in summary_df.columns and qc_col in summary_df.columns:
                summary_df['Bank ID'] = summary_df[id_col].astype(str).str.upper()
                quantile_map = dict(zip(summary_df['Bank ID'], summary_df[qc_col]))
                logger.info(f"Loaded quantile classifications from {summary_path}")
        except Exception as exc:
            logger.warning(f"Could not read {summary_path}: {exc}")

    if quantile_col not in df_components:
        df_components[quantile_col] = pd.Series(index=df_components.index, dtype=object)

    if quantile_map:
        df_components[quantile_col] = df_components['Bank ID'].map(quantile_map)

    thresholds = df_components['Overall Score'].quantile([0.25, 0.5, 0.75])

    def classify(score: float) -> str:
        if pd.isna(score):
            return "UNKNOWN"
        if score <= thresholds.loc[0.25]:
            return "MINIMAL"
        if score <= thresholds.loc[0.50]:
            return "LOW"
        if score <= thresholds.loc[0.75]:
            return "MEDIUM"
        return "HIGH"

    df_components[quantile_col] = df_components[quantile_col].fillna(
        df_components['Overall Score'].apply(classify)
    )

    ordered_cols = ['Bank ID', 'Overall Score', 'Risk Level', quantile_col]
    remaining_cols = [col for col in df_components.columns if col not in ordered_cols]
    return df_components[ordered_cols + remaining_cols]


def _format_component_sheet(worksheet):
    """Format the Component Scores sheet"""
    # Header formatting
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Column widths
    worksheet.column_dimensions['A'].width = 12   # Bank ID
    worksheet.column_dimensions['B'].width = 15   # Overall Score
    worksheet.column_dimensions['C'].width = 15   # Risk Level
    worksheet.column_dimensions['D'].width = 24   # Quantile Classification
    worksheet.column_dimensions['E'].width = 18   # Credit Risk Score
    worksheet.column_dimensions['F'].width = 20   # Liquidity Risk Score
    worksheet.column_dimensions['G'].width = 15   # Anomaly Score
    worksheet.column_dimensions['H'].width = 20   # Highest Risk Area
    
    # Freeze header row
    worksheet.freeze_panes = 'A2'


def _format_analysis_sheet(worksheet):
    """Format the Risk Analysis sheet"""
    # Header formatting
    header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Column widths
    worksheet.column_dimensions['A'].width = 12
    worksheet.column_dimensions['B'].width = 50
    worksheet.column_dimensions['C'].width = 50
    worksheet.column_dimensions['D'].width = 50
    worksheet.column_dimensions['E'].width = 60
    
    # Text wrapping for all cells
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    worksheet.freeze_panes = 'A2'


def _format_recommendations_sheet(worksheet):
    """Format the Recommendations sheet"""
    # Header formatting
    header_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Column widths
    worksheet.column_dimensions['A'].width = 12
    worksheet.column_dimensions['B'].width = 12
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 40
    worksheet.column_dimensions['E'].width = 60
    worksheet.column_dimensions['F'].width = 18
    
    # Text wrapping
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Color-code priorities
    for row in worksheet.iter_rows(min_row=2):
        priority_cell = row[1]  # Column B
        if priority_cell.value == 'CRITICAL':
            for cell in row:
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        elif priority_cell.value == 'HIGH':
            for cell in row:
                cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        elif priority_cell.value == 'MEDIUM':
            for cell in row:
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    
    worksheet.freeze_panes = 'A2'


# ============================================================================
# EXAMPLE 5: Custom Analysis Workflow
# ============================================================================

def run_custom_workflow_example():
    """
    Example of custom analysis workflow
    """
    logger.info("EXAMPLE 5: Custom Analysis Workflow")
    
    import importlib
    data_prep_module = importlib.import_module('1_data_preparation')
    anomaly_module = importlib.import_module('2_model_anomaly_detection')
    reporting_module = importlib.import_module('3_reporting_analysis')
    
    DataPreparation = data_prep_module.DataPreparation
    AnomalyDetectionModel = anomaly_module.AnomalyDetectionModel
    ReportingAnalysis = reporting_module.ReportingAnalysis
    
    # Load data
    df = pd.read_csv(DATASET_CSV)
    
    # Custom workflow: Focus on anomaly detection
    logger.info("Custom Workflow: Enhanced Anomaly Detection")
    
    # Prepare data
    data_prep = DataPreparation()
    data_prep.load_time_series_data(df, bank_id="VCB")
    features = data_prep.prepare_time_series_features()
    
    # Train anomaly detection with custom parameters
    anomaly_detector = AnomalyDetectionModel()
    training_results = anomaly_detector.train_models(
        features,
        contamination=0.05  # Expect 5% outliers
    )
    
    logger.info("Anomaly Detection Training Results:")
    for model, result in training_results.items():
        logger.debug(f"  - {model}: {result}")
    
    # Detect anomalies
    logger.info("Detecting anomalies...")
    anomaly_results = anomaly_detector.detect_anomalies(
        features,
        voting_threshold=0.4  # Lower threshold for more sensitivity
    )
    
    logger.info(f"  - Total Records: {anomaly_results['total_records']}")
    logger.info(f"  - Anomalies: {anomaly_results['anomalies_detected']}")
    logger.info(f"  - Anomaly Rate: {anomaly_results['anomaly_rate']:.2%}")
    
    # Detect fraud patterns
    logger.info("Detecting fraud patterns...")
    fraud_results = anomaly_detector.detect_fraud_patterns(features)
    
    logger.info(f"  - Patterns Detected: {fraud_results['patterns_detected']}")
    for pattern in fraud_results['patterns']:
        logger.debug(f"    • {pattern['pattern']}: {pattern['description']}")
    
    # Generate custom report
    reporting = ReportingAnalysis("ABC Bank", "2024")
    
    # You can create custom reports with selected components
    logger.info("Custom report generated successfully")


# ============================================================================
# EXAMPLE 6: Batch Processing (Optimized for Large Datasets)
# ============================================================================

def run_batch_audit_example():
    """
    Example of efficient batch processing for multiple banks using vectorized operations.
    Handles 100+ banks much faster than sequential per-bank loops.
    """
    logger.info("EXAMPLE 6: Batch Processing (Optimized)")
    
    try:
        from utils.batch_processing import BatchProcessor, BatchAuditRunner
        
        # Load sample data
        df = pd.read_csv(DATASET_CSV)
        logger.info(f"Loaded {len(df)} records for {df['bank_id'].nunique()} banks")
        
        # Initialize batch processor
        batch = BatchProcessor()
        
        # ===== BATCH OPERATION 1: Compute all banks' ratios at once =====
        logger.info("1. Computing ratios for all banks (vectorized)...")
        all_ratios = batch.prepare_all_banks_ratios(df)
        logger.info(f"   - Computed {len(all_ratios)} ratio records")
        
        # ===== BATCH OPERATION 2: Identify high-risk banks =====
        logger.info("2. Identifying high-risk banks (vectorized)...")
        # Load expert rule thresholds from config for batch identification
        try:
            from config.expert_rules_config import load_expert_rules
            rules = load_expert_rules()
            thresholds = {
                'npl_ratio': ('>', rules.get('asset_quality', {}).get('max_npl_ratio', 0.05)),
                'liquidity_ratio': ('<', rules.get('liquidity', {}).get('liquidity_ratio_min', 0.3)),
                'capital_adequacy_ratio': ('<', rules.get('capital_adequacy', {}).get('min_car', 0.08))
            }
        except Exception:
            thresholds = {
                'npl_ratio': ('>', 0.05),
                'liquidity_ratio': ('<', 0.3),
                'capital_adequacy_ratio': ('<', 0.08)
            }
        high_risk = batch.identify_high_risk_banks(all_ratios, thresholds)
        logger.info(f"   - Found {len(high_risk)} banks with violations")
        for bank_id, violations in list(high_risk.items())[:5]:
            logger.debug(f"     * {bank_id}: {', '.join(violations)}")
        
        # ===== BATCH OPERATION 3: Compute bank-level statistics =====
        logger.info("3. Computing statistics across all banks (vectorized)...")
        metrics = ['npl_ratio', 'liquidity_ratio', 'capital_adequacy_ratio', 'roa', 'roe']
        stats = batch.compute_bank_level_statistics(df, metrics)
        logger.info(f"   - Statistics for {len(stats)} banks")
        logger.debug(stats.head())
        
        # ===== BATCH OPERATION 4: Rank banks by metric =====
        logger.info("4. Ranking banks by NPL ratio (vectorized)...")
        ranks = batch.rank_banks_by_metric(df, 'npl_ratio', ascending=True)
        logger.info("   - Top 5 safest banks (lowest NPL):")
        logger.debug(ranks.head(5))
        
        # ===== BATCH OPERATION 5: Detect outlier banks =====
        logger.info("5. Detecting outlier banks (vectorized)...")
        outliers = batch.batch_outlier_detection(
            df,
            metrics=['loan_growth', 'asset_growth', 'roa'],
            z_score_threshold=2.5
        )
        logger.info(f"   - Found {len(outliers)} outlier banks")
        
        # ===== BATCH OPERATION 6: Run audits in batch =====
        logger.info("6. Running audits for all banks (batch mode)...")
        audit_runner = BatchAuditRunner(BankAuditSystem)
        bank_ids = df['bank_id'].unique()[:5]  # Limit to 5 for demo
        all_reports = audit_runner.run_batch_audits(
            df,
            bank_ids=bank_ids,
            audit_period='2024'
        )
        logger.info(f"   - Completed audits for {len(all_reports)} banks")
        
        # ===== BATCH OPERATION 7: Aggregate results =====
        logger.info("7. Aggregating batch results (vectorized)...")
        summary = audit_runner.aggregate_batch_results(all_reports)
        logger.debug(summary)
        
        logger.info("Batch processing completed successfully!")
        
        return all_reports
        
    except ImportError as e:
        logger.error(f"Import failed - {e}")
        return {}
    except Exception as e:
        logger.error(f"{type(e).__name__}: {e}")
        import traceback
        logger.debug(traceback.format_exc())
        return {}


# ============================================================================
# Main Execution
# ============================================================================

if __name__ == "__main__":
    logger.info("MODULAR BANK AUDIT SYSTEM - EXAMPLES")
    
    # Uncomment the example you want to run:
    
    # Example 1: Simple audit
    # report1 = run_simple_audit_example()
    
    # Example 2: Detailed audit
    # report2 = run_detailed_audit_example()
    
    # Example 3: Component-level usage
    # run_component_level_example()
    
    # Example 4: Multiple banks
    # all_reports = run_multiple_banks_example()
    
    # Example 5: Custom workflow
    # run_custom_workflow_example()
    
    # Example 6: Batch processing (OPTIMIZED)
    # run_batch_audit_example()
    
    # NEW: Audit ALL banks and save to outputs folder
    all_reports = audit_all_banks_to_outputs()
    
    # NEW: Export comprehensive component analysis to Excel
    print("\n")
    export_component_analysis_to_excel("outputs/banks_components_analysis.xlsx")
    
    # NEW: Generate quarterly risk scores
    print("\n" + "="*80)
    print("GENERATING QUARTERLY RISK SCORES...")
    print("="*80)
    try:
        # First generate the quarterly scores data
        quarterly_scores_gen.generate_quarterly_scores_for_all_banks()
        # Then create the formatted Excel file
        quarterly_generator.create_quarterly_results_excel()
        print("SUCCESS: Quarterly results generated successfully!")
    except Exception as e:
        print(f"Note: Quarterly generation encountered: {str(e)}")
        import traceback
        logger.debug(traceback.format_exc())
    
    print("\n" + "="*80)
    print("Examples completed!")
    print("="*80 + "\n")
